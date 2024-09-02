from datetime import datetime
import PySimpleGUI as sg
import pytesseract
import cv2
from openpyxl import load_workbook, Workbook
import os

# Function to crop the MRZ region from an image
def crop_mrz_region(image_path):
    """
    Crops the MRZ region from the image based on predefined coordinates.

    Parameters:
    - image_path (str): Path to the image file containing the MRZ.

    Returns:
    - cropped_image: Cropped image of the MRZ region.
    """
    img = cv2.imread(image_path)
    height, width = img.shape[:2]

    # Define the cropping region (you may need to adjust these values)
    mrz_region = img[int(height*0.7):height, 0:width]
    
    return mrz_region

# Function to extract MRZ text from an image
def extract_mrz_text(image_path):
    """
    Extracts MRZ text from an image file using PyTesseract.

    Parameters:
    - image_path (str): Path to the image file containing the MRZ.

    Returns:
    - str: Extracted MRZ text.
    """
    cropped_image = crop_mrz_region(image_path)
    mrz_text = pytesseract.image_to_string(cropped_image)
    
    # Clean the OCR output to remove unnecessary lines
    mrz_text = clean_ocr_output(mrz_text)
    
    # Apply OCR corrections
    mrz_text = correct_ocr_errors(mrz_text)
    
    return mrz_text

# Function to clean the OCR output by removing unnecessary lines at the top
def clean_ocr_output(text):
    """
    Cleans the OCR output by removing any unnecessary lines at the top,
    keeping only the last two lines, which typically contain the MRZ.

    Parameters:
    - text (str): The OCR text to be cleaned.

    Returns:
    - str: Cleaned OCR text containing only the MRZ.
    """
    lines = text.splitlines()
    return "\n".join(lines[-2:])

# Function to correct OCR errors
def correct_ocr_errors(text):
    """
    Corrects common OCR errors such as misinterpreting '<' as 'K', 'KS', or 'E'
    in specific areas of the MRZ text where '<' is expected.

    Parameters:
    - text (str): The OCR text to be corrected.

    Returns:
    - str: Corrected text.
    """
    corrected_text = []
    
    for line in text.splitlines():
        corrected_line = line.replace('K', '<').replace('KS', '<').replace('E', '<')
        corrected_text.append(corrected_line)
    
    return "\n".join(corrected_text)

# Function to manually parse MRZ data
def parse_mrz(mrz_text):
    lines = mrz_text.splitlines()
    if len(lines) < 2:
        raise ValueError("Invalid MRZ format")

    mrz_data = {}

    # First line of MRZ
    first_line = lines[0]
    mrz_data['surname'], mrz_data['given_names'] = first_line[5:].split('<<', 1)
    mrz_data['surname'] = clean_name(mrz_data['surname'])
    mrz_data['given_names'] = clean_name(mrz_data['given_names'])

    # Second line of MRZ
    second_line = lines[1]
    mrz_data['passport_number'] = second_line[0:9].replace('<', '')
    mrz_data['nationality'] = second_line[10:13]
    mrz_data['date_of_birth'] = format_date(second_line[13:19])
    mrz_data['sex'] = second_line[20]
    mrz_data['expiration_date'] = format_date(second_line[21:27])

    return mrz_data

# Function to clean up names by removing trailing 'K' or 'KS'
def clean_name(name):

    name = name.replace('<', ' ')
    
    if name.endswith('KS'):
        name = name[:-2]
    elif name.endswith('K'):
        name = name[:-1]
    
    return name.strip()

# Function to format date from YYMMDD to DD-MM-YYYY
def format_date(date_str):
    try:
        if len(date_str) == 6:
            date_obj = datetime.strptime(date_str, '%y%m%d')
            return date_obj.strftime('%d-%m-%Y')
        else:
            return ""
    except ValueError:
        return ""

# Function to update Excel with MRZ data
def update_excel_with_mrz(mrz_data, excel_path):
    try:
        if os.path.exists(excel_path):
            workbook = load_workbook(excel_path)
        else:
            workbook = Workbook()
        
        sheet = workbook.active
        
        next_row = sheet.max_row + 1
        
        if next_row == 2:
            sheet['A1'] = "Surname"
            sheet['B1'] = "Given Names"
            sheet['C1'] = "Passport Number"
            sheet['D1'] = "Nationality"
            sheet['E1'] = "Date of Birth"
            sheet['F1'] = "Sex"
            sheet['G1'] = "Date of Expiry"
        
        sheet[f'A{next_row}'] = mrz_data['surname']
        sheet[f'B{next_row}'] = mrz_data['given_names']
        sheet[f'C{next_row}'] = mrz_data['passport_number']
        sheet[f'D{next_row}'] = mrz_data['nationality']
        sheet[f'E{next_row}'] = mrz_data['date_of_birth']
        sheet[f'F{next_row}'] = mrz_data['sex']
        sheet[f'G{next_row}'] = mrz_data['expiration_date']
        
        workbook.save(excel_path)
        
        print("Excel sheet updated successfully!")
    except Exception as e:
        print(f"Error updating Excel file: {e}")

# PySimpleGUI layout
layout = [
    [sg.Text("Select Passport Image")],
    [sg.Input(), sg.FileBrowse(file_types=(("Image Files", "*.jpg;*.png;*.jpeg"),))],
    [sg.Button("Extract MRZ"), sg.Button("Cancel")],
    [sg.Text("", size=(40, 1), key="MRZText")],
    [sg.Button("Update Excel", disabled=True)]
]

window = sg.Window("MRZ Extraction and Excel Update", layout)

excel_path = 'passport_data.xlsx'

# Event loop for the PySimpleGUI window
while True:
    event, values = window.read()
    
    if event in (sg.WIN_CLOSED, 'Cancel'):
        break
    
    if event == 'Extract MRZ':
        image_path = values[0]
        if image_path:
            mrz_text = extract_mrz_text(image_path)
            window['MRZText'].update(mrz_text)
            window['Update Excel'].update(disabled=False)
    
    if event == 'Update Excel':
        mrz_text = window['MRZText'].get()
        if mrz_text:
            try:
                mrz_data = parse_mrz(mrz_text)
                update_excel_with_mrz(mrz_data, excel_path)
            except ValueError as e:
                print(f"Error parsing MRZ text: {e}")
            window['MRZText'].update("")
            window['Update Excel'].update(disabled=True)

window.close()
